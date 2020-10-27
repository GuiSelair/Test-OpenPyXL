# encoding: utf-8

import openpyxl

class OpenPyXLProvider:

    def openWorkbook(self, filename):
        try:
            self.workbook = openpyxl.load_workbook(filename) 
        except FileNotFoundError as err:
            print("[ERROR] > Erro ao encontrar o arquivo.")
        except Exception as err:
            print(err)

    def createWorkbook(self):
        try:
            self.workbook = openpyxl.Workbook() 
        except FileNotFoundError as err:
            print("[ERROR] > Erro ao encontrar o arquivo.")
        except PermissionError as err:
            print("[ERROR] > Verifique as permissões do diretório")
        except Exception as err:
            print(err)
    
    def saveWorkbook(self, filename):
        try:
            self.workbook.save(filename)
        except PermissionError as err:
            print("[ERROR] > Verifique as permissões do diretório")

    def listSheetsName(self):
        return self.workbook.sheetnames

    def openActiveSheet(self):
        self.sheet = self.workbook.active
        return self.sheet

    def openSheet(self, sheetName):
        self.sheet = self.workbook[sheetName]
        return self.sheet

